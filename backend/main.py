from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Cookie, Response, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.exceptions import RequestValidationError
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import timedelta
import uuid
import csv
import io
import datetime
import hashlib
import random
import os
from xhtml2pdf import pisa
from io import BytesIO
import qrcode
import base64
from dotenv import load_dotenv

import models, schemas, crypto_utils, database, auth_utils, oa_logic
import pdf_utils

load_dotenv()

# Create tables
models.Base.metadata.create_all(bind=database.engine)

app = FastAPI(title="EduCerts API")
templates = Jinja2Templates(directory="templates")

FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")
ENVIRONMENT = os.getenv("ENVIRONMENT", "development")

# CORS setup - strict origin
app.add_middleware(
    CORSMiddleware,
    allow_origins=[FRONTEND_URL],
    allow_credentials=True,  # Required for cookies
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    return JSONResponse(
        status_code=422,
        content={"detail": str(exc.errors()[0]["msg"]) if exc.errors() else "Validation error"}
    )

@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    import traceback
    print(f"GLOBAL ERROR: {exc}\n{traceback.format_exc()}")
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal server error"}
    )

def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

def normalize_column_name(header: str) -> str:
    """
    Normalizes a header name (lowercase, strip, underscores).
    Maps common aliases to student_name and course_name.
    """
    if not header:
        return ""
    h = str(header).lower().strip().replace(" ", "_").replace("-", "_")
    
    # Aliases for student_name
    if h in {"student_name", "student", "full_name", "name", "recipient", "recipient_name", "candidate_name", "student_fullname",
             "rollno", "roll_no", "enrn", "enrollment_no", "student_id", "reg_no", "registration_number"}:
        return "student_name"
    
    # Aliases for course_name
    if h in {"course_name", "course", "subject", "program", "training_name", "training", "module", "study_program",
             "subject_1", "subject_code", "course_code"}:
        return "course_name"

    return h

def get_current_user_from_cookie(
    access_token: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
) -> Optional[models.User]:
    """Dependency to extract and validate the user from the HttpOnly cookie."""
    if not access_token:
        return None
    payload = auth_utils.decode_access_token(access_token)
    if not payload:
        return None
    username = payload.get("sub")
    if not username:
        return None
    user = db.query(models.User).filter(models.User.name == username).first()
    return user

def require_user(current_user: Optional[models.User] = Depends(get_current_user_from_cookie)) -> models.User:
    """Dependency that requires an authenticated user."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return current_user

def require_admin(current_user: models.User = Depends(require_user)) -> models.User:
    """Dependency that requires an admin user."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

def generate_qr_base64(data: str):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode()

# ─────────────────────────────────────────────────────────────────────────────
# Auth Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/")
def read_root():
    return {"message": "EduCerts API — Secure Mode"}

@app.post("/api/signup")
def signup(user_data: schemas.UserCreate, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.email == user_data.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    if db.query(models.User).filter(models.User.name == user_data.name).first():
        raise HTTPException(status_code=400, detail="Name already taken")

    hashed_password = auth_utils.get_password_hash(user_data.password)
    new_user = models.User(
        name=user_data.name,
        email=user_data.email,
        password=hashed_password,
        is_admin=False
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"id": new_user.id, "name": new_user.name, "email": new_user.email}

@app.post("/api/login")
def login(response: Response, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(
        (models.User.name == form_data.username) | (models.User.email == form_data.username)
    ).first()

    if not user or not auth_utils.verify_password(form_data.password, user.password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
        )

    access_token_expires = timedelta(minutes=auth_utils.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth_utils.create_access_token(
        data={"sub": user.name}, expires_delta=access_token_expires
    )

    # Set token as secure HttpOnly cookie instead of returning it in body
    is_production = ENVIRONMENT == "production"
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,           # Not accessible via JavaScript
        secure=is_production,    # HTTPS-only in production
        samesite="lax",          # CSRF protection
        max_age=auth_utils.ACCESS_TOKEN_EXPIRE_MINUTES * 60
    )

    # Still return user info (but NOT the token)
    return {
        "user": {
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "is_admin": user.is_admin
        }
    }

@app.get("/api/me")
def get_current_user_info(current_user: models.User = Depends(require_user)):
    """Returns the currently authenticated user from the HttpOnly cookie."""
    return {
        "id": current_user.id,
        "name": current_user.name,
        "email": current_user.email,
        "is_admin": current_user.is_admin
    }

@app.post("/api/logout")
def logout(response: Response):
    """Clears the auth cookie."""
    response.delete_cookie(key="access_token")
    return {"message": "Logged out successfully"}

# ─────────────────────────────────────────────────────────────────────────────
# Certificate Issuance (Phase 3: Document Registry)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/issue", response_model=schemas.Certificate)
def issue_certificate(cert_data: schemas.CertificateCreate, db: Session = Depends(get_db)):
    cert_type = (cert_data.cert_type or "certificate").lower()

    # Build a generic document structure — no hardcoded transcript assumption
    raw_data = {
        "id": str(uuid.uuid4())[:8],
        "type": cert_type,
        "name": cert_data.course_name,
        "issuedOn": datetime.datetime.now().isoformat(),
        "recipient": {
            "name": cert_data.student_name,
            "studentId": cert_data.data_payload.get("student_id", "N/A")
        },
        # Spread all extra data_payload fields into the document (grade, department, gpa, etc.)
        **{k: v for k, v in cert_data.data_payload.items() if k not in ("student_id", "organization")}
    }

    organization = cert_data.data_payload.get("organization", "EduCerts Academy")
    issuers = [
        {
            "name": organization,
            "url": "https://educerts.io",
            "documentStore": "0x007d40224f6562461633ccfbaffd359ebb2fc9ba",
            "identityProof": {"type": "DNS-TXT", "location": "educerts.io"}
        }
    ]

    # Wrap and sign
    oa_doc = oa_logic.wrap_document(raw_data, issuers=issuers)
    merkle_root = oa_doc["signature"]["merkleRoot"]
    signature = crypto_utils.sign_data(merkle_root)
    oa_doc["signature"]["signature"] = signature
    oa_doc["signature"]["publicKey"] = crypto_utils.get_public_key_pem()

    # Anchor Merkle Root to Document Registry
    batch_id = str(uuid.uuid4())
    doc_registry_entry = models.DocumentRegistry(
        id=batch_id,
        merkle_root=merkle_root,
        issuer_name="EduCerts Admin",
        organization=organization,
        cert_count=1,
    )
    db.add(doc_registry_entry)

    claim_pin = "".join([str(random.randint(0, 9)) for _ in range(6)])

    cert_id = str(uuid.uuid4())
    db_cert = models.Certificate(
        id=cert_id,
        student_name=cert_data.student_name,
        course_name=cert_data.course_name,
        cert_type=cert_type,
        data_payload=oa_doc,
        signature=signature,
        claim_pin=claim_pin,
        organization=organization,
        batch_id=batch_id
    )
    db.add(db_cert)
    db.commit()
    db.refresh(db_cert)
    return db_cert


# ─────────────────────────────────────────────────────────────────────────────
# Claiming
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/claim")
def claim_certificate(claim_data: dict, db: Session = Depends(get_db)):
    pin = claim_data.get("pin")
    org = claim_data.get("organization")

    cert = db.query(models.Certificate).filter(
        models.Certificate.claim_pin == pin,
        models.Certificate.organization == org
    ).first()

    if not cert:
        raise HTTPException(status_code=404, detail="No certificate found for this PIN and Organization")
    if cert.revoked:
        raise HTTPException(status_code=400, detail="This certificate has been revoked")

    cert.claimed = True
    db.commit()
    return cert.data_payload

# ─────────────────────────────────────────────────────────────────────────────
# Verification (Phase 3: Check Document Registry)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/verify")
def verify_certificate(request: schemas.VerificationRequest, db: Session = Depends(get_db)):
    oa_doc = None
    cert = None

    if request.certificate_id:
        cert = db.query(models.Certificate).filter(models.Certificate.id == request.certificate_id).first()
        if not cert:
            raise HTTPException(status_code=404, detail="Certificate not found")
        oa_doc = cert.data_payload
    elif request.data_payload:
        oa_doc = request.data_payload
        signature = oa_doc.get("signature", {}).get("signature")
        cert = db.query(models.Certificate).filter(models.Certificate.signature == signature).first()

    if not oa_doc:
        raise HTTPException(status_code=400, detail="Must provide certificate_id or data_payload")

    # 1. Integrity Check
    merkle_root = oa_doc.get("signature", {}).get("merkleRoot")
    signature = oa_doc.get("signature", {}).get("signature")
    salted_data = oa_doc.get("data", {})
    field_hashes = oa_logic.get_field_hashes(salted_data)
    calculated_root = oa_logic.calculate_merkle_root(field_hashes)
    is_integrity_valid = (calculated_root == merkle_root)

    # 2. Document Status
    is_issued = cert is not None
    is_not_revoked = cert and not cert.revoked

    # 3. Issuer Identity
    issuer_name = "Unknown"
    is_identity_valid = False
    issuers_data = oa_doc.get("data", {}).get("issuers", {}).get("value", [])
    if issuers_data and issuers_data[0].get("name") in ["EduCerts Academy", cert.organization if cert else ""]:
        issuer_name = f"{issuers_data[0].get('name')} (Verified)"
        is_identity_valid = True

    # 4. Signature Check
    is_signature_valid = crypto_utils.verify_signature(merkle_root, signature) if signature and merkle_root else False

    # ── Phase 3: Document Registry Check ──
    is_registry_valid = False
    if merkle_root:
        registry_entry = db.query(models.DocumentRegistry).filter(
            models.DocumentRegistry.merkle_root == merkle_root,
            models.DocumentRegistry.revoked == False
        ).first()
        is_registry_valid = registry_entry is not None
    print(f"DEBUG VERIFY: Registry Valid: {is_registry_valid}")

    all_valid = is_integrity_valid and is_issued and is_not_revoked and is_identity_valid and is_signature_valid and is_registry_valid

    return {
        "summary": {
            "all": all_valid,
            "documentStatus": is_issued and is_not_revoked,
            "documentIntegrity": is_integrity_valid,
            "issuerIdentity": is_identity_valid,
            "signature": is_signature_valid,
            "registryCheck": is_registry_valid
        },
        "data": [
            {
                "type": "DOCUMENT_INTEGRITY",
                "name": "OpenAttestationHash",
                "data": is_integrity_valid,
                "status": "VALID" if is_integrity_valid else "INVALID"
            },
            {
                "type": "DOCUMENT_STATUS",
                "name": "OpenAttestationIssued",
                "data": {"issued": is_issued, "revoked": not is_not_revoked},
                "status": "VALID" if is_issued and is_not_revoked else "INVALID"
            },
            {
                "type": "ISSUER_IDENTITY",
                "name": "EduCertsRegistry",
                "status": "VALID" if is_identity_valid else "INVALID",
                "data": [{"name": issuer_name, "status": "VALID" if is_identity_valid else "INVALID"}]
            },
            {
                "type": "DOCUMENT_REGISTRY",
                "name": "MerkleRootAnchor",
                "data": {"merkle_root": merkle_root, "anchored": is_registry_valid},
                "status": "VALID" if is_registry_valid else "INVALID"
            }
        ],
        "certificate": {
            "student_name": cert.student_name if cert else "Unknown",
            "course_name": cert.course_name if cert else "Unknown"
        }
    }

# ─────────────────────────────────────────────────────────────────────────────
# Certificate CRUD
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/certificates", response_model=List[schemas.Certificate])
def get_all_certificates(db: Session = Depends(get_db)):
    return db.query(models.Certificate).order_by(models.Certificate.issued_at.desc()).all()

@app.get("/api/certificates/{student_name}", response_model=List[schemas.Certificate])
def get_student_certificates(student_name: str, db: Session = Depends(get_db)):
    return db.query(models.Certificate).filter(models.Certificate.student_name == student_name).all()

@app.post("/api/revoke/{cert_id}")
def revoke_certificate(cert_id: str, db: Session = Depends(get_db)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    cert.revoked = True
    # Also revoke the batch in Document Registry
    if cert.batch_id:
        registry = db.query(models.DocumentRegistry).filter(models.DocumentRegistry.id == cert.batch_id).first()
        if registry:
            registry.revoked = True
    db.commit()
    return {"message": "Certificate revoked and removed from Document Registry"}

# ─────────────────────────────────────────────────────────────────────────────
# Document Registry API
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/registry")
def get_document_registry(db: Session = Depends(get_db)):
    """Returns all anchored Merkle Roots — simulates querying the Document Store smart contract."""
    entries = db.query(models.DocumentRegistry).order_by(models.DocumentRegistry.anchored_at.desc()).all()
    return [
        {
            "id": e.id,
            "merkle_root": e.merkle_root,
            "issuer": e.issuer_name,
            "organization": e.organization,
            "cert_count": e.cert_count,
            "anchored_at": e.anchored_at.isoformat() if e.anchored_at else None,
            "revoked": e.revoked
        }
        for e in entries
    ]

# ─────────────────────────────────────────────────────────────────────────────
# Misc
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/apply/challenge")
def get_apply_challenge():
    return {
        "challenge": str(uuid.uuid4()),
        "required_fields": ["student_name", "course_name"],
        "message": "Scan with your Wallet App to verify your Degree."
    }

@app.post("/api/templates/upload")
async def upload_template(file: UploadFile = File(...)):
    if not file.filename.endswith(".html"):
        raise HTTPException(status_code=400, detail="Only .html files are allowed")
    content = await file.read()
    os.makedirs("user_templates", exist_ok=True)
    with open("user_templates/custom_certificate.html", "wb") as f:
        f.write(content)
    return {"message": "Template uploaded successfully", "template_name": file.filename}


# ─────────────────────────────────────────────────────────────────────────────
# PDF Template Upload & Parsing
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/templates/upload-pdf")
async def upload_pdf_template(file: UploadFile = File(...)):
    """
    Accept a PDF template file, extract all {{placeholder}} fields from its
    text layer, save the PDF for later use, and return the field list.
    """
    import re
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only .pdf files are allowed")

    content = await file.read()
    os.makedirs("user_templates", exist_ok=True)
    pdf_template_path = "user_templates/template.pdf"
    with open(pdf_template_path, "wb") as f:
        f.write(content)

    # Extract placeholders with their bounding boxes
    try:
        placeholder_map = pdf_utils.extract_pdf_placeholders(pdf_template_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse PDF: {e}")

    all_fields = list(placeholder_map.keys())
    system_fields = {"issued_at", "cert_id", "signature", "qr_code"}
    sig_fields = {"digital_signature", "stamp"}
    auto_fields = system_fields | sig_fields

    custom_fields = [f for f in all_fields if f not in auto_fields]
    input_fields = [f for f in all_fields if f not in {"issued_at", "cert_id", "signature", "qr_code"}]
    # Remove signature/stamp from user-facing input (they come from signer later)
    input_fields = [f for f in input_fields if f not in sig_fields]

    return {
        "all_fields": all_fields,
        "system_fields": [f for f in all_fields if f in system_fields],
        "signature_fields": [f for f in all_fields if f in sig_fields],
        "custom_fields": custom_fields,
        "input_fields": input_fields,
        "template_name": file.filename,
        "template_type": "pdf"
    }

@app.post("/api/templates/parse")
async def parse_template(file: UploadFile = File(...)):
    """
    Parses an uploaded HTML or PDF certificate template and extracts all {{placeholder}} fields.
    """
    import re
    filename_lower = file.filename.lower()

    # Redirect PDF uploads to the PDF parser
    if filename_lower.endswith(".pdf"):
        return await upload_pdf_template(file)

    if not filename_lower.endswith(".html"):
        raise HTTPException(status_code=400, detail="Only .html or .pdf files are allowed")

    content = (await file.read()).decode("utf-8", errors="ignore")

    os.makedirs("user_templates", exist_ok=True)
    with open("user_templates/custom_certificate.html", "w", encoding="utf-8") as f:
        f.write(content)

    placeholders = re.findall(r"\{\{\s*([\w\s]+?)\s*\}\}", content)
    seen = set()
    unique_fields = []
    for p in placeholders:
        p = p.strip()
        if p not in seen:
            seen.add(p)
            unique_fields.append(p)

    system_fields = {"student_name", "course_name", "issued_at", "cert_id", "signature", "qr_code"}
    sig_fields = {"digital_signature", "stamp"}
    custom_fields = [f for f in unique_fields if f not in system_fields]
    input_fields = [f for f in unique_fields if f not in {"issued_at", "cert_id", "signature", "qr_code"} and f not in sig_fields]

    return {
        "all_fields": unique_fields,
        "system_fields": [f for f in unique_fields if f in system_fields],
        "signature_fields": [f for f in unique_fields if f in sig_fields],
        "custom_fields": custom_fields,
        "input_fields": input_fields,
        "template_name": file.filename,
        "template_type": "html",
        "template_preview": content[:500] + "..." if len(content) > 500 else content
    }


@app.post("/api/templates/bulk-issue")
async def bulk_issue_from_template(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Reads a CSV file and issues one certificate per row.
    The previously uploaded template (user_templates/custom_certificate.html) is used.
    CSV column names are mapped directly to the template's {{ placeholder }} names.
    Required CSV columns: student_name, course_name (at minimum).
    """
    import re

    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")

    # Determine which template to use
    pdf_template_path = "user_templates/template.pdf"
    html_template_path = "user_templates/custom_certificate.html"
    use_pdf = os.path.exists(pdf_template_path)
    use_html = os.path.exists(html_template_path)

    if not use_pdf and not use_html:
        raise HTTPException(status_code=400, detail="No template uploaded yet.")

    template_path = pdf_template_path if use_pdf else html_template_path
    if use_pdf:
        # USE ROBUST PDF EXTRACTION
        placeholder_map = pdf_utils.extract_pdf_placeholders(template_path)
        template_fields = set(placeholder_map.keys())
    else:
        with open(template_path, "r", encoding="utf-8") as tf:
            template_text = tf.read()
        template_fields = set(re.findall(r"\{\{\s*([\w\s]+?)\s*\}\}", template_text))
        template_fields = {f.strip() for f in template_fields}

    # Parse the file
    content_bytes = await file.read()
    content = content_bytes.decode("utf-8", errors="ignore")
    csv_reader = csv.DictReader(io.StringIO(content))
    raw_rows = list(csv_reader)

    if not raw_rows:
        raise HTTPException(status_code=400, detail="CSV file is empty")

    rows = raw_rows
    headers = list(rows[0].keys())
    name_col = next((h for h in headers if normalize_column_name(h) == "student_name"), None)
    if not name_col:
        name_col = next((h for h in headers if "name" in h.lower() or "roll" in h.lower() or "id" in h.lower()), None)
        
    course_col = next((h for h in headers if normalize_column_name(h) == "course_name"), None)
    if not course_col:
        course_col = next((h for h in headers if "course" in h.lower() or "subject" in h.lower() or "prog" in h.lower()), None)
    
    issued_certs = []
    system_auto = {"issued_at", "cert_id", "signature", "qr_code", "digital_signature", "stamp"}
    os.makedirs("generated_certs", exist_ok=True)

    for row in rows:
        student_name = row.get(name_col, "").strip() if name_col else "Student"
        course_name = row.get(course_col, "").strip() if course_col else "Course"
        
        # Build data_payload: Match template placeholders to row keys (case-insensitive)
        data_payload_fields = {}
        row_keys_lower = {k.lower(): k for k in row.keys()}
        for field in template_fields:
            if field in system_auto: continue
            f_lower = field.lower()
            if f_lower in row_keys_lower:
                data_payload_fields[field] = row[row_keys_lower[f_lower]].strip()

        cert_type = row.get("cert_type", "certificate").strip() or "certificate"
        organization = row.get("organization", "EduCerts Academy").strip() or "EduCerts Academy"

        # Build raw OA document
        raw_data = {
            "id": str(uuid.uuid4())[:12],
            "type": cert_type,
            "name": course_name,
            "issuedOn": datetime.datetime.now().isoformat(),
            "recipient": {
                "name": student_name,
                "studentId": row.get("student_id", "N/A")
            },
            **{k: v for k, v in data_payload_fields.items() if k not in ("student_id", "organization")}
        }

        issuers = [{"name": organization, "url": "https://educerts.io",
                    "documentStore": "0x007d40224f6562461633ccfbaffd359ebb2fc9ba",
                    "identityProof": {"type": "DNS-TXT", "location": "educerts.io"}}]

        oa_doc = oa_logic.wrap_document(raw_data, issuers=issuers)
        merkle_root = oa_doc["signature"]["merkleRoot"]
        sig = crypto_utils.sign_data(merkle_root)
        oa_doc["signature"]["signature"] = sig
        oa_doc["signature"]["publicKey"] = crypto_utils.get_public_key_pem()

        batch_id = str(uuid.uuid4())
        db.add(models.DocumentRegistry(id=batch_id, merkle_root=merkle_root,
                                       issuer_name="EduCerts Admin", organization=organization, cert_count=1))

        claim_pin = "".join([str(random.randint(0, 9)) for _ in range(6)])
        cert_id = str(uuid.uuid4())

        # Render PDF if PDF template exists
        rendered_path = None
        if use_pdf:
            issued_at = datetime.datetime.now().strftime("%Y-%m-%d")
            field_values = {
                "student_name": student_name,
                "course_name": course_name,
                "issued_at": issued_at,
                "cert_id": cert_id,
                "signature": sig[:20] + "...",
                **data_payload_fields
            }
            out_path = f"generated_certs/{cert_id}_base.pdf"
            try:
                pdf_utils.render_pdf_certificate(pdf_template_path, field_values, out_path)
                rendered_path = out_path
            except Exception:
                rendered_path = None

        db_cert = models.Certificate(
            id=cert_id, student_name=student_name, course_name=course_name,
            cert_type=cert_type, data_payload=oa_doc, signature=sig,
            claim_pin=claim_pin, organization=organization, batch_id=batch_id,
            template_type="pdf" if use_pdf else "html",
            rendered_pdf_path=rendered_path,
            signing_status="unsigned"
        )
        db.add(db_cert)
        issued_certs.append({"id": cert_id, "student_name": student_name, "course_name": course_name, "signing_status": "unsigned"})

    db.commit()

    db.commit()
    return {
        "message": f"{len(issued_certs)} certificates issued from template",
        "count": len(issued_certs),
        "certificates": issued_certs
    }


@app.post("/api/templates/bulk-issue-excel")
async def bulk_issue_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Reads an Excel (.xlsx) OR CSV file and issues one certificate per row.
    Works the same as /api/templates/bulk-issue but supports Excel in addition to CSV.
    """
    import re
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are allowed")

    # Determine which template to use
    pdf_template_path = "user_templates/template.pdf"
    html_template_path = "user_templates/custom_certificate.html"
    use_pdf = os.path.exists(pdf_template_path)
    use_html = os.path.exists(html_template_path)

    if not use_pdf and not use_html:
        raise HTTPException(status_code=400, detail="No template uploaded. Upload a PDF or HTML template first.")

    template_path = pdf_template_path if use_pdf else html_template_path
    if use_pdf:
        # USE ROBUST PDF EXTRACTION
        placeholder_map = pdf_utils.extract_pdf_placeholders(template_path)
        template_fields = set(placeholder_map.keys())
    else:
        with open(template_path, "r", encoding="utf-8") as tf:
            template_text = tf.read()
        template_fields = set(re.findall(r"\{\{\s*([\w\s]+?)\s*\}\}", template_text))
        template_fields = {f.strip() for f in template_fields}

    # Parse the file
    content_bytes = await file.read()
    raw_rows = []
    if filename_lower.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content_bytes))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        content_str = content_bytes.decode("utf-8", errors="ignore")
        csv_reader = csv.DictReader(io.StringIO(content_str))
        raw_rows = list(csv_reader)

    if not raw_rows:
        raise HTTPException(status_code=400, detail="File is empty")
    
    rows = raw_rows
    headers = list(rows[0].keys())
    name_col = next((h for h in headers if normalize_column_name(h) == "student_name"), None)
    if not name_col:
        name_col = next((h for h in headers if "name" in h.lower() or "roll" in h.lower() or "id" in h.lower()), None)

    course_col = next((h for h in headers if normalize_column_name(h) == "course_name"), None)
    if not course_col:
        course_col = next((h for h in headers if "course" in h.lower() or "subject" in h.lower() or "prog" in h.lower() or "cent" in h.lower()), None)

    issued_certs = []
    system_auto = {"issued_at", "cert_id", "signature", "qr_code", "digital_signature", "stamp"}
    os.makedirs("generated_certs", exist_ok=True)

    for row in rows:
        student_name = row.get(name_col, "").strip() if name_col else "Student"
        course_name = row.get(course_col, "").strip() if course_col else "Course"
        
        # Build data_payload: Match template placeholders to row keys (case-insensitive)
        data_payload_fields = {}
        row_keys_lower = {k.lower(): k for k in row.keys()}
        for field in template_fields:
            if field in system_auto: continue
            f_lower = field.lower()
            if f_lower in row_keys_lower:
                data_payload_fields[field] = row[row_keys_lower[f_lower]].strip()

        cert_type = row.get("cert_type", "certificate").strip() or "certificate"
        organization = row.get("organization", "EduCerts Academy").strip() or "EduCerts Academy"

        raw_data = {
            "id": str(uuid.uuid4())[:8],
            "type": cert_type,
            "name": course_name,
            "issuedOn": datetime.datetime.now().isoformat(),
            "recipient": {"name": student_name, "studentId": row.get("student_id", "N/A")},
            **{k: v for k, v in data_payload_fields.items() if k not in ("student_id", "organization")}
        }
        issuers = [{"name": organization, "url": "https://educerts.io",
                    "documentStore": "0x007d40224f6562461633ccfbaffd359ebb2fc9ba",
                    "identityProof": {"type": "DNS-TXT", "location": "educerts.io"}}]

        oa_doc = oa_logic.wrap_document(raw_data, issuers=issuers)
        merkle_root = oa_doc["signature"]["merkleRoot"]
        sig = crypto_utils.sign_data(merkle_root)
        oa_doc["signature"]["signature"] = sig
        oa_doc["signature"]["publicKey"] = crypto_utils.get_public_key_pem()

        batch_id = str(uuid.uuid4())
        db.add(models.DocumentRegistry(id=batch_id, merkle_root=merkle_root,
                                       issuer_name="EduCerts Admin", organization=organization, cert_count=1))

        claim_pin = "".join([str(random.randint(0, 9)) for _ in range(6)])
        cert_id = str(uuid.uuid4())

        # Render PDF if PDF template exists
        rendered_path = None
        if use_pdf:
            issued_at = datetime.datetime.now().strftime("%Y-%m-%d")
            field_values = {
                "student_name": student_name,
                "course_name": course_name,
                "issued_at": issued_at,
                "cert_id": cert_id,
                "signature": sig[:20] + "...",
                **data_payload_fields
            }
            out_path = f"generated_certs/{cert_id}_base.pdf"
            try:
                print(f"DEBUG: Rendering PDF for cert {cert_id}...")
                pdf_utils.render_pdf_certificate(pdf_template_path, field_values, out_path)
                rendered_path = out_path
                print(f"DEBUG: PDF rendered successfully: {out_path}")
            except Exception as e:
                import traceback
                print(f"DEBUG: PDF RENDER ERROR: {e}")
                traceback.print_exc()
                rendered_path = None

        db_cert = models.Certificate(
            id=cert_id, student_name=student_name, course_name=course_name,
            cert_type=cert_type, data_payload=oa_doc, signature=sig,
            claim_pin=claim_pin, organization=organization, batch_id=batch_id,
            template_type="pdf" if use_pdf else "html",
            rendered_pdf_path=rendered_path,
            signing_status="unsigned"
        )
        db.add(db_cert)
        issued_certs.append({"id": cert_id, "student_name": student_name,
                              "course_name": course_name, "signing_status": "unsigned"})

    db.commit()
    return {
        "message": f"{len(issued_certs)} certificates issued",
        "count": len(issued_certs),
        "certificates": issued_certs
    }


# ─────────────────────────────────────────────────────────────────────────────
# Digital Signature Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/sign/upload")
async def upload_signature_assets(
    signature_file: Optional[UploadFile] = File(None),
    stamp_file: Optional[UploadFile] = File(None),
    signer_name: str = Form(...),
    signer_role: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Upload a digital signature image and/or stamp image.
    Stores them under user_templates/ and saves a DigitalSignatureRecord.
    """
    os.makedirs("user_templates", exist_ok=True)
    sig_path = None
    stamp_path = None

    if signature_file and signature_file.filename:
        sig_bytes = await signature_file.read()
        sig_path = f"user_templates/signature_{current_user.id}.png"
        with open(sig_path, "wb") as f:
            f.write(sig_bytes)

    if stamp_file and stamp_file.filename:
        stamp_bytes = await stamp_file.read()
        stamp_path = f"user_templates/stamp_{current_user.id}.png"
        with open(stamp_path, "wb") as f:
            f.write(stamp_bytes)

    record = models.DigitalSignatureRecord(
        signer_name=signer_name,
        signer_role=signer_role,
        signature_path=sig_path,
        stamp_path=stamp_path
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return {
        "id": record.id,
        "signer_name": signer_name,
        "signer_role": signer_role,
        "signature_uploaded": sig_path is not None,
        "stamp_uploaded": stamp_path is not None
    }


@app.post("/api/sign/apply")
async def apply_digital_signatures(
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Apply digital signature/stamp to one or multiple certificates.
    Body: {
      cert_ids: [...],           # list of cert IDs to sign
      signer_name: str,
      signer_role: str,
      signature_record_id: int   # optional — use a previously uploaded record
    }
    OR if no record_id, looks for the current user's latest uploaded sig.
    """
    cert_ids = body.get("cert_ids", [])
    signer_name = body.get("signer_name", current_user.name)
    signer_role = body.get("signer_role", "Authorized Signatory")
    record_id = body.get("signature_record_id")

    # Find signature record
    if record_id:
        sig_record = db.query(models.DigitalSignatureRecord).filter(
            models.DigitalSignatureRecord.id == record_id
        ).first()
    else:
        sig_record = db.query(models.DigitalSignatureRecord).order_by(
            models.DigitalSignatureRecord.uploaded_at.desc()
        ).first()

    sig_path = sig_record.signature_path if sig_record else None
    stamp_path = sig_record.stamp_path if sig_record else None

    pdf_template_path = "user_templates/template.pdf"
    has_pdf_template = os.path.exists(pdf_template_path)

    os.makedirs("generated_certs", exist_ok=True)
    signed_certs = []
    now_iso = datetime.datetime.now().isoformat()

    for cert_id in cert_ids:
        cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
        if not cert:
            continue

        signed_pdf_path = f"generated_certs/{cert_id}_signed.pdf"

        if cert.template_type == "pdf" and has_pdf_template:
            base_path = cert.rendered_pdf_path or pdf_template_path
            if not os.path.exists(base_path):
                base_path = pdf_template_path
            try:
                pdf_utils.apply_signatures_to_pdf(
                    pdf_path=base_path,
                    signature_img_path=sig_path,
                    stamp_img_path=stamp_path,
                    template_path=pdf_template_path,
                    output_path=signed_pdf_path
                )
                cert.rendered_pdf_path = signed_pdf_path
            except Exception as e:
                print(f"Signing error for {cert_id}: {e}")
                continue
        else:
            # HTML-based cert — re-render with signature embedded
            verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
            qr_b64 = generate_qr_base64(verify_url)
            html_template_path = "user_templates/custom_certificate.html"
            if os.path.exists(html_template_path):
                from jinja2 import FileSystemLoader, Environment
                env = Environment(loader=FileSystemLoader("user_templates"))
                tmpl = env.get_template("custom_certificate.html")
            else:
                tmpl = templates.get_template("certificate.html")

            render_ctx = {
                "student_name": cert.student_name,
                "course_name": cert.course_name,
                "issued_at": cert.issued_at.strftime("%Y-%m-%d"),
                "cert_id": cert.id,
                "signature": cert.signature[:30] + "...",
                "qr_code": qr_b64,
            }
            html_content = tmpl.render(**render_ctx)
            result = BytesIO()
            pisa.pisaDocument(BytesIO(html_content.encode("utf-8")), result)
            with open(signed_pdf_path, "wb") as f:
                f.write(result.getvalue())
            cert.rendered_pdf_path = signed_pdf_path

        # Update signing metadata
        existing_sigs = cert.digital_signatures or []
        existing_sigs.append({
            "signer_name": signer_name,
            "signer_role": signer_role,
            "applied_at": now_iso
        })
        cert.digital_signatures = existing_sigs
        cert.signing_status = "signed"
        signed_certs.append({"id": cert_id, "student_name": cert.student_name})

    db.commit()
    return {
        "message": f"{len(signed_certs)} certificates signed",
        "signed": signed_certs
    }


@app.post("/api/sign/apply-batch/{batch_id}")
async def apply_signatures_to_batch(
    batch_id: str,
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Apply digital signature to all certs in a batch."""
    certs = db.query(models.Certificate).filter(
        models.Certificate.batch_id == batch_id
    ).all()
    cert_ids = [c.id for c in certs]
    body["cert_ids"] = cert_ids
    # Delegate to apply endpoint logic
    return await apply_digital_signatures(body, db, current_user)


@app.get("/api/certificates/unsigned")
def get_unsigned_certificates(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Returns all certificates that have not yet been digitally signed."""
    certs = db.query(models.Certificate).filter(
        models.Certificate.signing_status == "unsigned",
        models.Certificate.revoked == False
    ).order_by(models.Certificate.issued_at.desc()).all()
    return [
        {
            "id": c.id,
            "student_name": c.student_name,
            "course_name": c.course_name,
            "cert_type": c.cert_type,
            "issued_at": c.issued_at.isoformat() if c.issued_at else None,
            "organization": c.organization,
            "signing_status": c.signing_status,
            "template_type": c.template_type
        }
        for c in certs
    ]


@app.get("/api/sign/records")
def get_signature_records(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Return all uploaded signature/stamp records."""
    records = db.query(models.DigitalSignatureRecord).order_by(
        models.DigitalSignatureRecord.uploaded_at.desc()
    ).all()
    return [
        {
            "id": r.id,
            "signer_name": r.signer_name,
            "signer_role": r.signer_role,
            "has_signature": r.signature_path is not None and os.path.exists(r.signature_path or ""),
            "has_stamp": r.stamp_path is not None and os.path.exists(r.stamp_path or ""),
            "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else None
        }
        for r in records
    ]


@app.post("/api/import")
async def import_data(file: UploadFile = File(...)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")
    content = await file.read()
    csv_reader = csv.DictReader(io.StringIO(content.decode('utf-8')))
    data = [row for row in csv_reader if "student_name" in row and "course_name" in row]
    return {"message": "Data imported successfully", "count": len(data), "data": data}

@app.get("/api/download/{cert_id}")
def download_certificate(cert_id: str, db: Session = Depends(get_db)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")

    # ── If a rendered/signed PDF already exists, serve it directly ──
    if cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path):
        return FileResponse(
            path=cert.rendered_pdf_path,
            media_type="application/pdf",
            filename=f"cert_{cert.id}.pdf"
        )

    # ── PDF template path ──
    pdf_template_path = "user_templates/template.pdf"
    if cert.template_type == "pdf" and os.path.exists(pdf_template_path):
        # Render on-the-fly from PDF template
        verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
        qr_b64 = generate_qr_base64(verify_url)
        field_values = {
            "student_name": cert.student_name,
            "course_name": cert.course_name,
            "issued_at": cert.issued_at.strftime("%Y-%m-%d"),
            "cert_id": cert.id,
            "signature": cert.signature[:20] + "...",
            "qr_code": qr_b64,
        }
        # Also overlay payload fields - ROBUST EXTRACTION
        payload_data = cert.data_payload or {}
        # Try both direct payload and OA 'data' nested payload
        candidates = [payload_data, payload_data.get("data", {})]
        
        for source in candidates:
            if not isinstance(source, dict): continue
            for k, v in source.items():
                # OA might have values like {"value": "John"} or just "John"}
                if isinstance(v, dict) and "value" in v:
                    field_values.setdefault(k, v["value"])
                elif isinstance(v, (str, int, float)):
                    field_values.setdefault(k, v)
                elif isinstance(v, dict):
                    # Check nested objects like recipient.name
                    for subk, subv in v.items():
                        if isinstance(subv, (str, int, float)):
                            field_values.setdefault(f"{k}_{subk}", subv)
                            field_values.setdefault(subk, subv)

        print(f"DEBUG: On-the-fly field values: {list(field_values.keys())}")
        os.makedirs("generated_certs", exist_ok=True)
        out_path = f"generated_certs/{cert.id}_base.pdf"
        try:
            pdf_utils.render_pdf_certificate(pdf_template_path, field_values, out_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"PDF render error: {e}")

        # Save the rendered path for next time
        cert.rendered_pdf_path = out_path
        db.commit()

        return FileResponse(
            path=out_path,
            media_type="application/pdf",
            filename=f"cert_{cert.id}.pdf"
        )

    # ── Fallback: HTML template → xhtml2pdf ──
    # ONLY if it's not a PDF template. If it's a PDF template and we are here, something is wrong.
    if cert.template_type == "pdf":
        raise HTTPException(status_code=500, detail="PDF template was requested but rendering failed or template is missing.")

    verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
    qr_base64 = generate_qr_base64(verify_url)

    custom_template_path = "user_templates/custom_certificate.html"
    if os.path.exists(custom_template_path):
        from jinja2 import FileSystemLoader, Environment
        env = Environment(loader=FileSystemLoader("user_templates"))
        template = env.get_template("custom_certificate.html")
    else:
        template = templates.get_template("certificate.html")

    render_ctx = {
        "student_name": cert.student_name,
        "course_name": cert.course_name,
        "issued_at": cert.issued_at.strftime("%Y-%m-%d"),
        "cert_id": cert.id,
        "signature": cert.signature[:30] + "...",
        "qr_code": qr_base64,
    }

    payload_data = cert.data_payload or {}
    extra_fields = {}
    for k, v in payload_data.items():
        if k not in ("signature", "data", "schema") and isinstance(v, (str, int, float)):
            extra_fields[k] = v
    oa_data = payload_data.get("data", {})
    if isinstance(oa_data, dict):
        for k, v in oa_data.items():
            if isinstance(v, dict) and "value" in v:
                extra_fields[k] = v["value"]
            elif isinstance(v, (str, int, float)):
                extra_fields[k] = v

    render_ctx = {**extra_fields, **render_ctx}
    html_content = template.render(**render_ctx)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("utf-8")), result)
    if pdf.err:
        raise HTTPException(status_code=500, detail="Error generating PDF")

    return Response(
        content=result.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cert_{cert.id}.pdf"}
    )

@app.get("/api/json/{cert_id}")
def download_json_certificate(cert_id: str, db: Session = Depends(get_db)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    return JSONResponse(
        content=cert.data_payload,
        headers={"Content-Disposition": f"attachment; filename=cert_{cert.id}.json"}
    )
